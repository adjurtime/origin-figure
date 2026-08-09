#!/usr/bin/env python3
"""Privacy-safe, read-only structural audit for Origin figure artifacts."""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import sys
import xml.etree.ElementTree as ET
from contextlib import suppress
from pathlib import Path
from typing import Any
from zipfile import BadZipFile


SCHEMA_VERSION = "2.0"
DEFAULT_MAX_CELLS = 1_000_000
DEFAULT_MAX_FILE_SIZE_MB = 100.0

GRAPHIC_TAGS = {
    "circle",
    "ellipse",
    "g",
    "image",
    "line",
    "path",
    "polygon",
    "polyline",
    "rect",
    "text",
    "use",
}

VECTOR_TAGS = {
    "circle",
    "ellipse",
    "line",
    "path",
    "polygon",
    "polyline",
    "rect",
    "use",
}


class OptionalDependencyError(RuntimeError):
    """Raised when a requested inspection needs an uninstalled optional package."""


class ResourceLimitError(ValueError):
    """Raised when an input exceeds the configured bounded-inspection limits."""


INSPECTION_ERRORS = (
    OSError,
    ValueError,
    KeyError,
    BadZipFile,
    ET.ParseError,
    csv.Error,
    OptionalDependencyError,
)


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def display_path(path: Path, full_paths: bool) -> str:
    return str(path) if full_paths else path.name


def file_record(path: Path, full_paths: bool) -> dict[str, Any]:
    return {
        "path": display_path(path, full_paths),
        "size_bytes": path.stat().st_size,
        "sha256": sha256(path),
    }


def json_safe(value: Any) -> Any:
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    if hasattr(value, "isoformat"):
        with suppress(TypeError, ValueError):
            return value.isoformat()
    return str(value)


def load_openpyxl() -> tuple[Any, type[Exception]]:
    try:
        from openpyxl import load_workbook
        from openpyxl.utils.exceptions import InvalidFileException
    except ModuleNotFoundError as exc:
        raise OptionalDependencyError(
            "XLSX/XLSM inspection requires the optional dependency openpyxl; "
            "CSV, SVG, OPJU, and reference inspection do not require it"
        ) from exc
    return load_workbook, InvalidFileException


def inspect_workbook(
    path: Path,
    sample_limit: int,
    max_cells: int,
    include_preview: bool,
    include_formulas: bool,
    full_paths: bool,
) -> dict[str, Any]:
    load_workbook, invalid_file_error = load_openpyxl()
    formulas = None
    values = None
    try:
        try:
            formulas = load_workbook(
                path, read_only=False, data_only=False, keep_links=False
            )
            values = load_workbook(
                path, read_only=False, data_only=True, keep_links=False
            )
        except invalid_file_error as exc:
            raise ValueError(f"invalid workbook: {exc}") from exc

        declared_cells = sum(
            sheet.max_row * sheet.max_column for sheet in formulas.worksheets
        )
        if declared_cells > max_cells:
            raise ResourceLimitError(
                f"workbook declares {declared_cells} cells, exceeding "
                f"--max-cells={max_cells}"
            )

        sheets: list[dict[str, Any]] = []
        disclose_coordinates = include_preview or include_formulas

        for formula_sheet in formulas.worksheets:
            value_sheet = values[formula_sheet.title]
            formula_count = 0
            missing_cached_results = 0
            nonempty_count = 0
            formula_samples: list[dict[str, Any]] = []
            missing_cached_coordinates: list[str] = []
            header_preview: list[list[Any]] = []

            if include_preview:
                preview_columns = min(formula_sheet.max_column, sample_limit)
                for row in formula_sheet.iter_rows(
                    min_row=1,
                    max_row=min(formula_sheet.max_row, 3),
                    min_col=1,
                    max_col=preview_columns,
                ):
                    header_preview.append([json_safe(cell.value) for cell in row])

            for row in formula_sheet.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        nonempty_count += 1
                    if cell.data_type == "f" or (
                        isinstance(cell.value, str) and cell.value.startswith("=")
                    ):
                        formula_count += 1
                        cached_value = value_sheet[cell.coordinate].value
                        if include_formulas and len(formula_samples) < sample_limit:
                            sample: dict[str, Any] = {
                                "coordinate": cell.coordinate,
                                "formula": str(cell.value),
                            }
                            if include_preview:
                                sample["cached_value"] = json_safe(cached_value)
                            formula_samples.append(sample)
                        if cached_value is None:
                            missing_cached_results += 1
                            if (
                                disclose_coordinates
                                and len(missing_cached_coordinates) < sample_limit
                            ):
                                missing_cached_coordinates.append(cell.coordinate)

            cell_area = formula_sheet.max_row * formula_sheet.max_column
            sheet_record: dict[str, Any] = {
                "name": formula_sheet.title,
                "dimension": formula_sheet.calculate_dimension(),
                "max_row": formula_sheet.max_row,
                "max_column": formula_sheet.max_column,
                "declared_cells": cell_area,
                "nonempty_cells": nonempty_count,
                "empty_cells_in_declared_area": max(cell_area - nonempty_count, 0),
                "formula_cells": formula_count,
                "formula_cells_without_cached_result": missing_cached_results,
                "merged_ranges": len(formula_sheet.merged_cells.ranges),
            }
            if include_preview:
                sheet_record["header_preview_first_3_rows"] = header_preview
            if include_formulas:
                sheet_record["formula_samples"] = formula_samples
            if disclose_coordinates:
                sheet_record["missing_cached_result_coordinates"] = (
                    missing_cached_coordinates
                )
            sheets.append(sheet_record)

        return {
            **file_record(path, full_paths),
            "declared_cells": declared_cells,
            "sheets": sheets,
        }
    finally:
        if formulas is not None:
            formulas.close()
        if values is not None:
            values.close()


def inspect_csv(
    path: Path,
    sample_limit: int,
    max_cells: int,
    include_preview: bool,
    full_paths: bool,
) -> dict[str, Any]:
    encoding = None
    sample_text = None
    for candidate in ("utf-8-sig", "utf-8", "gb18030"):
        try:
            with path.open("r", encoding=candidate, newline="") as handle:
                sample_text = handle.read(65536)
            encoding = candidate
            break
        except UnicodeDecodeError:
            continue
    if encoding is None or sample_text is None:
        raise ValueError("CSV encoding is not UTF-8 or GB18030")

    try:
        dialect = csv.Sniffer().sniff(sample_text, delimiters=",\t;|")
    except csv.Error:
        dialect = csv.excel

    row_count = 0
    max_columns = 0
    parsed_cells = 0
    blank_cells = 0
    preview: list[list[str]] = []
    with path.open("r", encoding=encoding, newline="") as handle:
        for row in csv.reader(handle, dialect):
            row_count += 1
            parsed_cells += len(row)
            if parsed_cells > max_cells:
                raise ResourceLimitError(
                    f"CSV exceeds --max-cells={max_cells} during inspection"
                )
            max_columns = max(max_columns, len(row))
            blank_cells += sum(not value.strip() for value in row)
            if include_preview and len(preview) < sample_limit:
                preview.append(row[:sample_limit])

    record: dict[str, Any] = {
        **file_record(path, full_paths),
        "encoding": encoding,
        "delimiter": dialect.delimiter,
        "rows": row_count,
        "max_columns": max_columns,
        "parsed_cells": parsed_cells,
        "blank_cells": blank_cells,
    }
    if include_preview:
        record["preview"] = preview
    return record


def local_name(tag: str) -> str:
    return tag.rsplit("}", 1)[-1]


def inspect_svg(path: Path, full_paths: bool) -> dict[str, Any]:
    root = ET.parse(path).getroot()
    if local_name(root.tag) != "svg":
        raise ValueError("root element is not <svg>")
    graphic_elements = sum(
        1 for element in root.iter() if local_name(element.tag) in GRAPHIC_TAGS
    )
    vector_elements = sum(
        1 for element in root.iter() if local_name(element.tag) in VECTOR_TAGS
    )
    embedded_images = sum(
        1 for element in root.iter() if local_name(element.tag) == "image"
    )
    text_elements = sum(
        1 for element in root.iter() if local_name(element.tag) == "text"
    )
    return {
        **file_record(path, full_paths),
        "width": root.attrib.get("width"),
        "height": root.attrib.get("height"),
        "viewBox": root.attrib.get("viewBox"),
        "graphic_elements": graphic_elements,
        "vector_elements": vector_elements,
        "embedded_images": embedded_images,
        "text_elements": text_elements,
    }


def existing_file(
    path_text: str,
    kind: str,
    issues: list[str],
    max_bytes: int,
    full_paths: bool,
) -> Path | None:
    unresolved = Path(path_text).expanduser()
    unresolved_label = str(unresolved) if full_paths else unresolved.name
    try:
        path = unresolved.resolve()
        label = display_path(path, full_paths)
        if not path.is_file():
            issues.append(f"{kind} file not found: {label}")
            return None
        size_bytes = path.stat().st_size
        if size_bytes == 0:
            issues.append(f"{kind} file is empty: {label}")
            return None
        if size_bytes > max_bytes:
            issues.append(
                f"{kind} file exceeds --max-file-size-mb: {label} "
                f"({size_bytes} bytes)"
            )
            return None
        return path
    except OSError as exc:
        detail = str(exc) if full_paths else type(exc).__name__
        issues.append(f"cannot access {kind} file {unresolved_label}: {detail}")
        return None


def positive_int(value: str) -> int:
    parsed = int(value)
    if parsed < 1:
        raise argparse.ArgumentTypeError("must be at least 1")
    return parsed


def positive_float(value: str) -> float:
    parsed = float(value)
    if parsed <= 0:
        raise argparse.ArgumentTypeError("must be greater than 0")
    return parsed


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Inspect workbook, CSV, OPJU, SVG, and reference artifacts without "
            "modifying them. Output is privacy-safe by default."
        )
    )
    parser.add_argument(
        "--xlsx",
        action="append",
        default=[],
        help="Source XLSX or XLSM path (requires optional openpyxl)",
    )
    parser.add_argument("--csv", action="append", default=[], help="Source CSV path")
    parser.add_argument("--svg", action="append", default=[], help="SVG path")
    parser.add_argument("--opju", action="append", default=[], help="OPJU path")
    parser.add_argument(
        "--reference",
        action="append",
        default=[],
        help="Other source or visual reference to hash",
    )
    parser.add_argument(
        "--sample-limit",
        type=positive_int,
        default=10,
        help="Maximum preview, formula, and coordinate items (default: 10)",
    )
    parser.add_argument(
        "--max-cells",
        type=positive_int,
        default=DEFAULT_MAX_CELLS,
        help=(
            "Maximum declared workbook cells or parsed CSV cells "
            f"(default: {DEFAULT_MAX_CELLS})"
        ),
    )
    parser.add_argument(
        "--max-file-size-mb",
        type=positive_float,
        default=DEFAULT_MAX_FILE_SIZE_MB,
        help=(
            "Maximum input size in MiB "
            f"(default: {DEFAULT_MAX_FILE_SIZE_MB:g})"
        ),
    )
    parser.add_argument(
        "--include-preview",
        action="store_true",
        help=(
            "Include bounded workbook headers and CSV rows; when formulas are "
            "included, also include cached values"
        ),
    )
    parser.add_argument(
        "--include-formulas",
        action="store_true",
        help="Include bounded formula text and cell coordinates",
    )
    parser.add_argument(
        "--full-paths",
        action="store_true",
        help="Include resolved absolute paths instead of basenames",
    )
    parser.add_argument(
        "--require-outputs",
        action="store_true",
        help="Require at least one valid OPJU and one valid SVG",
    )
    return parser.parse_args()


def main() -> int:
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8")
    args = parse_args()
    max_bytes = int(args.max_file_size_mb * 1024 * 1024)
    privacy_safe = not (
        args.include_preview or args.include_formulas or args.full_paths
    )
    issues: list[str] = []
    result: dict[str, Any] = {
        "schema_version": SCHEMA_VERSION,
        "scope": "file_structure_and_bounded_tabular_profile",
        "privacy_mode": "safe" if privacy_safe else "expanded",
        "disclosure": {
            "full_paths": args.full_paths,
            "previews": args.include_preview,
            "formulas": args.include_formulas,
        },
        "limits": {
            "max_cells": args.max_cells,
            "max_file_size_mb": args.max_file_size_mb,
            "sample_limit": args.sample_limit,
        },
        "limitations": [
            "OPJU checks cover only existence, non-empty size, and SHA-256.",
            "Reopen, editability, graph objects, and data bindings require Origin MCP verification.",
            "Workbook and CSV scans stop when configured resource limits are exceeded.",
            "Header, row, formula, cached-value, coordinate, and full-path disclosure is opt-in.",
        ],
        "workbooks": [],
        "csv_files": [],
        "svgs": [],
        "opju_files": [],
        "references": [],
    }

    if not (args.xlsx or args.csv or args.svg or args.opju or args.reference):
        issues.append("no artifacts were supplied")

    def resolve(path_text: str, kind: str) -> Path | None:
        return existing_file(
            path_text, kind, issues, max_bytes, args.full_paths
        )

    for path_text in args.xlsx:
        path = resolve(path_text, "XLSX/XLSM")
        if path is None:
            continue
        try:
            result["workbooks"].append(
                inspect_workbook(
                    path,
                    args.sample_limit,
                    args.max_cells,
                    args.include_preview,
                    args.include_formulas,
                    args.full_paths,
                )
            )
        except INSPECTION_ERRORS as exc:
            issues.append(
                f"cannot inspect XLSX/XLSM {display_path(path, args.full_paths)}: {exc}"
            )

    for path_text in args.csv:
        path = resolve(path_text, "CSV")
        if path is None:
            continue
        try:
            result["csv_files"].append(
                inspect_csv(
                    path,
                    args.sample_limit,
                    args.max_cells,
                    args.include_preview,
                    args.full_paths,
                )
            )
        except INSPECTION_ERRORS as exc:
            issues.append(
                f"cannot inspect CSV {display_path(path, args.full_paths)}: {exc}"
            )

    for path_text in args.svg:
        path = resolve(path_text, "SVG")
        if path is None:
            continue
        try:
            svg_record = inspect_svg(path, args.full_paths)
            result["svgs"].append(svg_record)
            label = display_path(path, args.full_paths)
            if svg_record["graphic_elements"] == 0:
                issues.append(f"SVG has no graphical elements: {label}")
            if not (
                svg_record["viewBox"]
                or (svg_record["width"] and svg_record["height"])
            ):
                issues.append(f"SVG has neither viewBox nor width/height: {label}")
            if svg_record["embedded_images"] and not svg_record["vector_elements"]:
                issues.append(f"SVG contains images but no vector geometry: {label}")
        except INSPECTION_ERRORS as exc:
            issues.append(
                f"cannot inspect SVG {display_path(path, args.full_paths)}: {exc}"
            )

    for path_text in args.opju:
        path = resolve(path_text, "OPJU")
        if path is not None:
            result["opju_files"].append(
                {
                    **file_record(path, args.full_paths),
                    "validation_scope": "existence_size_sha256_only",
                }
            )

    for path_text in args.reference:
        path = resolve(path_text, "reference")
        if path is not None:
            result["references"].append(
                {
                    **file_record(path, args.full_paths),
                    "suffix": path.suffix.lower(),
                }
            )

    if args.require_outputs:
        if not result["opju_files"]:
            issues.append("no valid OPJU output was supplied")
        if not result["svgs"]:
            issues.append("no valid SVG output was supplied")

    result["ok"] = not issues
    result["issues"] = issues
    json.dump(result, sys.stdout, ensure_ascii=False, indent=2)
    sys.stdout.write("\n")
    return 0 if result["ok"] else 1


if __name__ == "__main__":
    raise SystemExit(main())
